FILE_NAME = "NMC Nurse Scrapes_sheet 1.xlsx"
sheet_name = "Extract 1_ NMC List View"

import os
import json
current_directory = os.getcwd()
PATH = current_directory + "/" + FILE_NAME



import openpyxl
from tqdm import tqdm
from time import time
s = time()
print("loading workbook ...")

# Load the Excel file into an openpyxl workbook object
workbook = openpyxl.load_workbook(PATH)

print("loading workbook finished... in ", time() - s)


worksheet = workbook[sheet_name]

max_rows = worksheet.max_row
hyperlinks = set()

print("reading hyperlinks ...")

for i in  tqdm(range(2,max_rows)):
    try:
        hyperlinks.add(worksheet.cell(row=i, column=6).hyperlink.target)
    except Exception as e:
        print(e)
        print(f"Failed to get Details page from row {i}")


with open("hyperlinks_1.json", "w") as f:
    f.write(json.dumps(list(hyperlinks)))
print("reading hyperlinks finished ... in ", time()- s)
